"""
Metadata extractor for Office documents.
"""

import os
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional
import xml.etree.ElementTree as ET

try:
    from docx import Document
    from openpyxl import load_workbook
    from pptx import Presentation
except ImportError:
    Document = None
    load_workbook = None
    Presentation = None

try:
    import olefile
    from oletools.olevba import VBA_Parser
except ImportError:
    olefile = None
    VBA_Parser = None

from ..models import FileMetadata


class MetadataExtractor:
    """Extracts metadata from Office documents."""

    def __init__(self):
        """Initialize the metadata extractor."""
        pass

    def extract(self, file_path: str) -> FileMetadata:
        """
        Extract metadata from an Office file.

        Args:
            file_path: Path to the Office file

        Returns:
            FileMetadata object containing extracted metadata
        """
        file_path = Path(file_path)
        metadata = FileMetadata()

        try:
            # Get file timestamps
            stat = file_path.stat()
            metadata.creation_time = datetime.fromtimestamp(stat.st_ctime)
            metadata.modified_time = datetime.fromtimestamp(stat.st_mtime)

            # Check if it's a modern Office format (ZIP-based)
            if self._is_modern_format(file_path):
                self._extract_modern_metadata(str(file_path), metadata)
            else:
                self._extract_legacy_metadata(str(file_path), metadata)

        except Exception as e:
            # If extraction fails, return basic metadata
            pass

        return metadata

    def _is_modern_format(self, file_path: Path) -> bool:
        """Check if file is a modern Office format (ZIP-based)."""
        modern_extensions = {".docx", ".xlsx", ".pptx", ".docm", ".xlsm", ".pptm"}
        return file_path.suffix.lower() in modern_extensions

    def _extract_modern_metadata(self, file_path: str, metadata: FileMetadata) -> None:
        """Extract metadata from modern Office formats."""
        try:
            # Try to extract using specific library based on file type
            ext = Path(file_path).suffix.lower()

            if ext in [".docx", ".docm"] and Document:
                self._extract_word_metadata(file_path, metadata)
            elif ext in [".xlsx", ".xlsm"] and load_workbook:
                self._extract_excel_metadata(file_path, metadata)
            elif ext in [".pptx", ".pptm"] and Presentation:
                self._extract_powerpoint_metadata(file_path, metadata)
            else:
                # Fallback to ZIP-based extraction
                self._extract_zip_metadata(file_path, metadata)

        except Exception:
            # Fallback to ZIP-based extraction
            self._extract_zip_metadata(file_path, metadata)

    def _extract_word_metadata(self, file_path: str, metadata: FileMetadata) -> None:
        """Extract metadata from Word documents."""
        try:
            doc = Document(file_path)
            core_props = doc.core_properties

            metadata.title = core_props.title or None
            metadata.subject = core_props.subject or None
            metadata.author = core_props.author or None
            metadata.language = core_props.language or None
            metadata.last_saved_by = core_props.last_modified_by or None

            # Check for password protection
            metadata.password_protected = self._is_password_protected(file_path)

        except Exception:
            pass

    def _extract_excel_metadata(self, file_path: str, metadata: FileMetadata) -> None:
        """Extract metadata from Excel workbooks."""
        try:
            wb = load_workbook(file_path, read_only=True)

            if hasattr(wb, "properties"):
                props = wb.properties
                metadata.title = props.title or None
                metadata.subject = props.subject or None
                metadata.author = props.creator or None
                metadata.company = props.company or None
                metadata.last_saved_by = props.lastModifiedBy or None

            # Check for password protection
            metadata.password_protected = self._is_password_protected(file_path)

            # Check for hidden sheets
            hidden_sheets = [sheet.title for sheet in wb.worksheets if sheet.sheet_state != "visible"]
            if hidden_sheets:
                metadata.embedded_files = True

        except Exception:
            pass

    def _extract_powerpoint_metadata(self, file_path: str, metadata: FileMetadata) -> None:
        """Extract metadata from PowerPoint presentations."""
        try:
            prs = Presentation(file_path)
            core_props = prs.core_properties

            metadata.title = core_props.title or None
            metadata.subject = core_props.subject or None
            metadata.author = core_props.author or None
            metadata.last_saved_by = core_props.last_modified_by or None

            # Check for password protection
            metadata.password_protected = self._is_password_protected(file_path)

        except Exception:
            pass

    def _extract_zip_metadata(self, file_path: str, metadata: FileMetadata) -> None:
        """Extract metadata from ZIP-based Office files using raw XML parsing."""
        try:
            with zipfile.ZipFile(file_path, "r") as zip_file:
                # Try to read core properties
                try:
                    core_xml = zip_file.read("docProps/core.xml")
                    self._parse_core_properties(core_xml, metadata)
                except KeyError:
                    pass

                # Try to read app properties
                try:
                    app_xml = zip_file.read("docProps/app.xml")
                    self._parse_app_properties(app_xml, metadata)
                except KeyError:
                    pass

        except Exception:
            pass

    def _parse_core_properties(self, xml_content: bytes, metadata: FileMetadata) -> None:
        """Parse core properties XML."""
        try:
            root = ET.fromstring(xml_content)

            # Define namespaces
            namespaces = {
                "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
                "dc": "http://purl.org/dc/elements/1.1/",
                "dcterms": "http://purl.org/dc/terms/",
            }

            # Extract properties
            title_elem = root.find(".//dc:title", namespaces)
            if title_elem is not None:
                metadata.title = title_elem.text

            subject_elem = root.find(".//dc:subject", namespaces)
            if subject_elem is not None:
                metadata.subject = subject_elem.text

            creator_elem = root.find(".//dc:creator", namespaces)
            if creator_elem is not None:
                metadata.author = creator_elem.text

            language_elem = root.find(".//dc:language", namespaces)
            if language_elem is not None:
                metadata.language = language_elem.text

            modified_by_elem = root.find(".//cp:lastModifiedBy", namespaces)
            if modified_by_elem is not None:
                metadata.last_saved_by = modified_by_elem.text

        except Exception:
            pass

    def _parse_app_properties(self, xml_content: bytes, metadata: FileMetadata) -> None:
        """Parse app properties XML."""
        try:
            root = ET.fromstring(xml_content)

            # Define namespace
            namespace = {"": "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"}

            # Extract properties
            company_elem = root.find(".//Company", namespace)
            if company_elem is not None:
                metadata.company = company_elem.text

            manager_elem = root.find(".//Manager", namespace)
            if manager_elem is not None:
                metadata.manager = manager_elem.text

            template_elem = root.find(".//Template", namespace)
            if template_elem is not None:
                metadata.template = template_elem.text

            version_elem = root.find(".//AppVersion", namespace)
            if version_elem is not None:
                metadata.office_version = version_elem.text

        except Exception:
            pass

    def _extract_legacy_metadata(self, file_path: str, metadata: FileMetadata) -> None:
        """Extract metadata from legacy Office formats."""
        if not olefile:
            return

        try:
            if olefile.isOleFile(file_path):
                ole = olefile.OleFileIO(file_path)

                # Extract OLE properties
                if ole.exists("\\x05SummaryInformation"):
                    props = ole.getproperties("\\x05SummaryInformation")

                    # Map OLE property IDs to metadata fields
                    if 2 in props:  # Title
                        metadata.title = props[2]
                    if 3 in props:  # Subject
                        metadata.subject = props[3]
                    if 4 in props:  # Author
                        metadata.author = props[4]
                    if 8 in props:  # Last saved by
                        metadata.last_saved_by = props[8]
                    if 15 in props:  # Company
                        metadata.company = props[15]

                ole.close()

        except Exception:
            pass

    def _is_password_protected(self, file_path: str) -> bool:
        """Check if the document is password protected."""
        try:
            # For modern formats, try to open as ZIP
            if self._is_modern_format(Path(file_path)):
                try:
                    with zipfile.ZipFile(file_path, "r") as zip_file:
                        # Check for encrypted content
                        if "EncryptedPackage" in zip_file.namelist():
                            return True
                        if any("encrypt" in name.lower() for name in zip_file.namelist()):
                            return True
                except zipfile.BadZipFile:
                    # If we can't open as ZIP, it might be encrypted
                    return True
            else:
                # For legacy formats, check OLE structure
                if olefile and olefile.isOleFile(file_path):
                    ole = olefile.OleFileIO(file_path)
                    streams = ole.listdir()
                    ole.close()

                    # Look for encryption indicators
                    for stream in streams:
                        if any("encrypt" in part.lower() for part in stream):
                            return True
        except Exception:
            pass

        return False
